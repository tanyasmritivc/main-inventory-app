
import logging

from fastapi import APIRouter, Depends, File, Form, HTTPException, Request, UploadFile
from openai import OpenAI

from app.core.auth import AuthenticatedUser, get_current_user
from app.core.config import get_settings
from app.core.limiter import limiter
from app.services import sharing_service
from app.services.documents_repo import create_activity
from app.services.items_repo import bulk_create_items, list_items
from app.services.limits import TeamSoftCapExceeded, check_and_increment_import
from app.services.spaces_repo import SpaceLimitExceeded
from app.services.usage_service import check_limit, increment_usage

router = APIRouter(tags=["inventory"])

logger = logging.getLogger(__name__)


def _resolve_import_target(
    *, requesting_user_id: str, location: str, share_id: str | None
) -> tuple[str, str]:
    if not share_id:
        return requesting_user_id, location

    try:
        share, can_edit = sharing_service.get_share_access(
            requesting_user_id=requesting_user_id,
            share_id=share_id,
        )
    except ValueError as exc:
        raise HTTPException(403, str(exc)) from exc

    if not can_edit:
        raise HTTPException(403, 'You only have view access to this space.')

    share_name = (share.get('share_name') or '').strip()
    owner_user_id = (share.get('owner_user_id') or '').strip()
    if not share_name or not owner_user_id:
        raise HTTPException(422, 'This shared space is not available for imports.')
    return owner_user_id, share_name


def _resolve_analysis_target(
    *, requesting_user_id: str, location: str, share_id: str | None
) -> tuple[str, str]:
    if not share_id:
        return requesting_user_id, location
    try:
        share, _ = sharing_service.get_share_access(
            requesting_user_id=requesting_user_id,
            share_id=share_id,
        )
    except ValueError as exc:
        raise HTTPException(403, str(exc)) from exc
    share_name = (share.get('share_name') or '').strip()
    owner_user_id = (share.get('owner_user_id') or '').strip()
    if not share_name or not owner_user_id:
        raise HTTPException(422, 'This shared space is not available for analysis.')
    return owner_user_id, share_name


def _is_date_like(s: str) -> bool:
    import re
    s = str(s).strip()
    # Matches patterns like 2023-03-04, 2023-03-04 00:00:00, etc.
    return bool(re.match(
        r'^\d{4}-\d{2}-\d{2}', s))


def _normalized_identifier(value: object) -> str:
    import re
    return re.sub(r'[^a-z0-9]', '', str(value or '').lower())


def _column_index(headers: list, candidates: tuple[str, ...]) -> int | None:
    normalized = [_normalized_identifier(header) for header in headers]
    for candidate in candidates:
        key = _normalized_identifier(candidate)
        if key in normalized:
            return normalized.index(key)
    return None


def _parse_bom_rows(*, raw: bytes, filename: str) -> list[dict]:
    """Parse common BOM columns deterministically; analysis must not require AI."""
    import csv
    import io
    import openpyxl

    sheets: list[tuple[list, list]] = []
    if filename.endswith('.xlsx'):
        try:
            workbook = openpyxl.load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
            for worksheet in workbook.worksheets:
                rows = [list(row) for row in worksheet.iter_rows(values_only=True)]
                rows = [row for row in rows if any(str(cell or '').strip() for cell in row)]
                if len(rows) >= 2:
                    sheets.append((rows[0], rows[1:]))
            workbook.close()
        except Exception as exc:
            raise HTTPException(422, 'This Excel file could not be read.') from exc
    elif filename.endswith('.csv'):
        try:
            rows = list(csv.reader(io.StringIO(raw.decode('utf-8-sig', errors='replace'))))
            rows = [row for row in rows if any(str(cell or '').strip() for cell in row)]
            if len(rows) >= 2:
                sheets.append((rows[0], rows[1:]))
        except Exception as exc:
            raise HTTPException(422, 'This CSV file could not be read.') from exc
    else:
        raise HTTPException(400, 'Choose an Excel (.xlsx) or CSV file.')

    parsed: list[dict] = []
    for headers, rows in sheets:
        name_idx = _column_index(headers, ('name', 'item name', 'description', 'part description', 'product'))
        part_idx = _column_index(headers, ('part number', 'part #', 'part no', 'sku', 'item number', 'item #', 'product code'))
        qty_idx = _column_index(headers, ('quantity', 'qty', 'count', 'required', 'amount'))
        brand_idx = _column_index(headers, ('brand', 'manufacturer', 'vendor', 'supplier'))
        if name_idx is None and part_idx is None:
            continue
        for row in rows:
            def value(index: int | None) -> str:
                return str(row[index] or '').strip() if index is not None and index < len(row) else ''

            name, part_number = value(name_idx), value(part_idx)
            if not name and not part_number:
                continue
            try:
                required = max(1, int(float(value(qty_idx) or '1')))
            except (TypeError, ValueError):
                required = 1
            parsed.append({
                'name': name or part_number,
                'part_number': part_number or None,
                'brand': value(brand_idx) or None,
                'required_quantity': required,
            })
    if not parsed:
        raise HTTPException(
            422,
            'No BOM rows were found. Include a Name or Part Number column and a Quantity column.',
        )
    return parsed


@router.post('/inventory/bom/analyze')
@limiter.limit('5/minute')
async def analyze_bom_route(
    request: Request,
    file: UploadFile = File(...),
    location: str = Form('Unsorted'),
    share_id: str | None = Form(default=None),
    user: AuthenticatedUser = Depends(get_current_user),
):
    target_user_id, target_location = _resolve_analysis_target(
        requesting_user_id=user.user_id,
        location=location,
        share_id=share_id,
    )
    raw = await file.read()
    if not raw:
        raise HTTPException(400, 'The selected file is empty.')
    if len(raw) > 10 * 1024 * 1024:
        raise HTTPException(413, 'Choose a spreadsheet smaller than 10 MB.')

    bom_rows = _parse_bom_rows(raw=raw, filename=(file.filename or '').lower())
    inventory = [
        item for item in list_items(user_id=target_user_id)
        if str(item.get('location') or 'Unsorted').strip().lower() == target_location.strip().lower()
    ]

    results = []
    for bom in bom_rows:
        part_key = _normalized_identifier(bom.get('part_number'))
        name_key = _normalized_identifier(bom.get('name'))
        matches = []
        for item in inventory:
            item_part = _normalized_identifier(item.get('part_number'))
            item_name = _normalized_identifier(item.get('name'))
            if (part_key and item_part == part_key) or (not part_key and name_key and item_name == name_key):
                matches.append(item)
        available = sum(max(0, int(item.get('quantity') or 0)) for item in matches)
        required = bom['required_quantity']
        missing = max(0, required - available)
        results.append({
            **bom,
            'available_quantity': available,
            'missing_quantity': missing,
            'status': 'ready' if missing == 0 else ('partial' if available > 0 else 'missing'),
        })

    ready = sum(1 for row in results if row['status'] == 'ready')
    partial = sum(1 for row in results if row['status'] == 'partial')
    missing = sum(1 for row in results if row['status'] == 'missing')
    return {
        'location': target_location,
        'summary': {
            'total_lines': len(results),
            'ready_lines': ready,
            'partial_lines': partial,
            'missing_lines': missing,
            'readiness_percent': round(ready * 100 / len(results)),
        },
        'items': results,
    }


@router.post('/import/spreadsheet')
@limiter.limit("5/minute")
async def import_spreadsheet_route(
    request: Request,
    file: UploadFile = File(...),
    location: str = Form('Unsorted'),
    share_id: str | None = Form(default=None),
    user: AuthenticatedUser = Depends(get_current_user),
):
    import io as _io
    import csv as _csv
    import json as _json
    import re as _re
    import openpyxl

    target_user_id, target_location = _resolve_import_target(
        requesting_user_id=user.user_id,
        location=location,
        share_id=share_id,
    )

    raw = await file.read()
    if not raw:
        raise HTTPException(400, 'Empty file')

    filename = (file.filename or '').lower()
    all_sheets = []

    if filename.endswith('.xlsx') or filename.endswith('.xls'):
        try:
            wb = openpyxl.load_workbook(
                _io.BytesIO(raw), read_only=True, data_only=True)
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                rows = []
                for row in ws.iter_rows(values_only=True):
                    cleaned = [str(c).strip() if c is not None else '' for c in row]
                    if any(c for c in cleaned):
                        rows.append(cleaned)
                if len(rows) >= 2:
                    all_sheets.append({'name': sheet_name, 'headers': rows[0], 'rows': rows[1:]})
            wb.close()
        except Exception as e:
            raise HTTPException(422, f'Cannot read Excel: {e}')

    elif filename.endswith('.csv'):
        try:
            text = raw.decode('utf-8', errors='replace')
            reader = _csv.DictReader(_io.StringIO(text))
            headers = list(reader.fieldnames or [])
            rows = [[str(row.get(h, '')).strip() for h in headers] for row in reader]
            if rows:
                all_sheets.append({'name': 'Sheet1', 'headers': headers, 'rows': rows})
        except Exception as e:
            raise HTTPException(422, f'Cannot read CSV: {e}')
    else:
        raise HTTPException(400, 'Only .xlsx .xls .csv supported')

    if not all_sheets:
        raise HTTPException(422, 'No data found')

    # Build compact structure for AI — NOT all rows
    structure_lines = []
    for sheet in all_sheets:
        structure_lines.append(f"Sheet: {sheet['name']}")
        structure_lines.append(f"Columns: {', '.join(str(h) for h in sheet['headers'])}")
        structure_lines.append(f"Row count: {len(sheet['rows'])}")
        for i, row in enumerate(sheet['rows'][:3]):
            pairs = ', '.join(
                f'{sheet["headers"][j]}="{v}"'
                for j, v in enumerate(row)
                if v and j < len(sheet['headers'])
            )
            structure_lines.append(f"Sample {i+1}: {pairs}")
        structure_lines.append('')

    structure = '\n'.join(structure_lines)
    first_sheet = all_sheets[0]
    columns = ', '.join(str(h) for h in first_sheet['headers'])
    sample_rows_list = []
    for _sr_idx, _sr_row in enumerate(first_sheet['rows'][:3]):
        sample_rows_list.append({
            str(first_sheet['headers'][j]): _sr_row[j]
            for j in range(min(len(first_sheet['headers']), len(_sr_row)))
            if _sr_row[j]
        })
    sample_rows = _json.dumps(sample_rows_list, ensure_ascii=False)

    mapping_prompt = f"""You are mapping spreadsheet columns to inventory item fields.

Available inventory fields:
- name_columns: list of columns that form the item name/description (required)
- quantity_column: column with numeric quantity/count
- category_column: column with item type, category, or classification (e.g. "Category", "Type", "Class", "Group")
- part_number_column: column with part numbers, SKUs, item codes, IDs (e.g. "PN-F177", "SKU", "Part #", "Item #", column named with a number like "8")
- subcategory_column: column with size, type, dimension, shaft size, screw length, thread size (e.g. "Shaft", "Screw Length", "Size", "M4", "Thread")
- brand_column: column with vendor name, supplier, manufacturer, brand (e.g. "Vendor Name", "Supplier", "Brand", "Manufacturer")
- purchase_source_column: column with vendor part numbers, supplier codes, order numbers
- notes_columns: list of columns with additional info, descriptions, specifications, comments

Spreadsheet structure:
Columns: {columns}
Sample rows: {sample_rows}

Rules:
- category_column: any column named "category", "type", "class", "group", or similar classification -> map it here; return null if no such column exists
- part_number_column: if a column name is just a number (like "8") or contains "part", "PN", "SKU", "ID", "code" -> map it here
- subcategory_column: shaft size, screw length, thread size, M2/M3/M4/M5/M6/M8 values -> map here
- brand_column: anything with "vendor", "supplier", "manufacturer", "brand" -> map here
- name_columns: use Description field if exists, otherwise combine meaningful text columns
- notes_columns: any remaining descriptive columns not mapped elsewhere
- Never use date columns as names
- quantity_column: only one column, must contain numbers

Return ONLY valid JSON, no markdown, no explanation:
{{
  "name_columns": ["col1"],
  "quantity_column": "col2",
  "category_column": "col3",
  "part_number_column": "col4",
  "subcategory_column": "col5",
  "brand_column": "col6",
  "purchase_source_column": "col7",
  "notes_columns": ["col8"],
  "category": "Supplies",
  "display_columns": [
    {{"field": "name", "label": "Name"}},
    {{"field": "part_number", "label": "Part #"}},
    {{"field": "subcategory", "label": "Size/Type"}},
    {{"field": "brand", "label": "Vendor"}},
    {{"field": "purchase_source", "label": "Vendor Part #"}},
    {{"field": "quantity", "label": "Qty"}},
    {{"field": "notes", "label": "Notes"}}
  ]
}}
Only include fields in display_columns that actually have data in this spreadsheet.
Always include name and quantity."""

    settings = get_settings()
    ai_client = OpenAI(api_key=settings.openai_api_key)

    try:
        resp = ai_client.chat.completions.create(
            model='gpt-4o',
            max_completion_tokens=600,
            messages=[{'role': 'user', 'content': mapping_prompt}])
        mapping_raw = resp.choices[0].message.content.strip()
        if '```' in mapping_raw:
            for part in mapping_raw.split('```'):
                p = part.strip()
                if p.startswith('json'):
                    mapping_raw = p[4:].strip(); break
                elif p.startswith('{'):
                    mapping_raw = p; break
        mapping = _json.loads(mapping_raw)
    except Exception:
        mapping = {'name_columns': [], 'quantity_column': 'Quantity',
                   'category_column': None,
                   'part_number_column': None, 'subcategory_column': None,
                   'brand_column': None, 'purchase_source_column': None,
                   'notes_columns': [], 'category': 'Supplies'}

    def _get_val(row_dict: dict, col: str) -> str:
        if not col:
            return ''
        for k, v in row_dict.items():
            if str(k).strip().lower() == str(col).strip().lower():
                s = str(v).strip()
                return '' if s.lower() in ('none', 'nan', 'null', '') else s
        return ''

    items_to_insert = []
    for sheet in all_sheets:
        headers = sheet['headers']
        for row in sheet['rows']:
            row_dict = {
                headers[i]: row[i]
                for i in range(min(len(headers), len(row)))
            }

            name_columns = mapping.get('name_columns') or []
            if isinstance(name_columns, str):
                name_columns = [name_columns]

            notes_columns = mapping.get('notes_columns') or []
            if isinstance(notes_columns, str):
                notes_columns = [notes_columns]

            name = ' - '.join([
                _get_val(row_dict, col)
                for col in name_columns
                if _get_val(row_dict, col)
            ])
            if not name:
                continue

            qty_col = mapping.get('quantity_column') or ''
            qty_str = _get_val(row_dict, qty_col)
            try:
                qty = max(1, int(float(qty_str))) if qty_str else 1
            except Exception:
                qty = 1

            category = (
                _get_val(row_dict, mapping.get('category_column') or '')
                or mapping.get('category')
                or 'Supplies'
            )

            item = {
                'name': name[:500],
                'category': category,
                'subcategory': _get_val(row_dict, mapping.get('subcategory_column') or '') or None,
                'quantity': qty,
                'location': target_location,
                'part_number': _get_val(row_dict, mapping.get('part_number_column') or '') or None,
                'brand': _get_val(row_dict, mapping.get('brand_column') or '') or None,
                'purchase_source': _get_val(row_dict, mapping.get('purchase_source_column') or '') or None,
                'notes': ' | '.join([
                    _get_val(row_dict, col)
                    for col in notes_columns
                    if _get_val(row_dict, col)
                ]) or None,
            }
            items_to_insert.append(item)

    if not items_to_insert:
        raise HTTPException(422, 'No valid items found')

    # Team plan: check+increment team-level import counter; personal: use usage_limits.
    try:
        check_and_increment_import(user.user_id)
    except TeamSoftCapExceeded as exc:
        raise HTTPException(
            status_code=403,
            detail={
                'error': 'TEAM_SOFT_CAP',
                'feature': exc.feature,
                'current': exc.current,
                'limit': exc.limit,
                'resets_at': exc.resets_at,
                'message': f"Your team has used {exc.current} of {exc.limit} {exc.feature} for this period.",
            },
        )

    # Personal plan (check_and_increment_import returned without raising).
    limit_check = await check_limit(user.user_id, 'spreadsheet_import')
    if not limit_check['allowed']:
        raise HTTPException(
            status_code=403,
            detail={
                'error': 'FREE_TIER_IMPORT_LIMIT',
                'used': limit_check['current'],
                'max': limit_check['limit'],
            },
        )
    await increment_usage(user.user_id, 'spreadsheet_import')

    try:
        inserted, failures = bulk_create_items(user_id=target_user_id, items=items_to_insert)
    except SpaceLimitExceeded:
        raise HTTPException(403, "FREE_TIER_SPACE_LIMIT")

    try:
        create_activity(
            user_id=target_user_id,
            summary=f'Imported {len(inserted)} items from {file.filename}',
            metadata={
                'type': 'spreadsheet_import',
                'filename': file.filename,
                'inserted': len(inserted),
                'failures': len(failures),
                'location': target_location,
                'requested_by': user.user_id,
            },
        )
    except Exception:
        logger.warning('Spreadsheet import activity logging failed', exc_info=True)

    return {
        'inserted': len(inserted),
        'failures': len(failures),
        'total_found': len(items_to_insert),
        'display_columns': mapping.get('display_columns', [
            {'field': 'name', 'label': 'Name'},
            {'field': 'quantity', 'label': 'Qty'},
            {'field': 'location', 'label': 'Location'},
        ]),
    }
