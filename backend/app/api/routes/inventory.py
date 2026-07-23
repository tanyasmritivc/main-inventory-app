from __future__ import annotations

from fastapi import APIRouter, Depends, File, Form, HTTPException, Request, Response, UploadFile
from fastapi import status
from fastapi.responses import StreamingResponse
import logging

from typing import Any

from pydantic import BaseModel

import io

import httpx
import anyio
import openai
from openai import OpenAI
from PIL import Image

from app.core.auth import AuthenticatedUser, get_current_user
from app.core.config import get_settings
from app.core.errors import bad_gateway, bad_request, service_unavailable
from app.schemas.ai import AICommandRequest, AICommandResponse
from app.schemas.inventory import (
    AddItemRequest,
    AddItemResponse,
    DeleteItemResponse,
    ExtractFromImageResponse,
    ProcessBarcodeRequest,
    ProcessBarcodeResponse,
    BarcodeLookupRequest,
    BarcodeLookupResponse,
    SearchItemsRequest,
    SearchItemsResponse,
    UpdateItemRequest,
    UpdateItemResponse,
    BulkCreateRequest,
    BulkCreateResponse,
    MultiExtractFromImageResponse,
)
from app.schemas.documents import ListDocumentsResponse, RecentActivityResponse, UploadDocumentResponse
from app.services.items_repo import add_item, bulk_create_items, check_free_tier_limits, delete_item, search_items_basic, update_item
from app.services import sharing_service
from app.services.ai_agent import iter_ai_command_sse, run_ai_command
from app.services.openai_service import (
    extract_item_from_image,
    extract_items_from_image_multi,
    interpret_barcode,
    iter_assist_file_analysis_sse,
    parse_search_query_to_keywords,
    summarize_activity,
)
from app.services.documents_repo import (
    create_activity,
    create_document,
    list_documents,
    list_recent_activity,
    rename_document,
    set_document_item_link,
)
from app.services.supabase_client import get_supabase_admin
from app.services.catalog_service import lookup_in_catalog, save_to_catalog
from app.services.storage import upload_document, upload_image
from app.services.usage_service import (
    check_item_limit,
    FREE_ITEM_LIMIT,
    FREE_SCAN_LIMIT,
    is_pro_user,
    check_limit,
    increment_usage,
    get_all_usage,
)

def _convert_to_jpeg(image_bytes: bytes, filename: str) -> tuple[bytes, str]:
    if filename.lower().endswith(".heic"):
        try:
            img = Image.open(io.BytesIO(image_bytes))
            output = io.BytesIO()
            img.convert("RGB").save(output, format="JPEG", quality=85)
            return output.getvalue(), "converted.jpg"
        except Exception:
            pass
    return image_bytes, filename


router = APIRouter(tags=["inventory"])


logger = logging.getLogger(__name__)


_INVALID_NAMES = {
    "unknown",
    "unknown item",
    "n/a",
    "na",
    "none",
    "null",
}


def is_valid_result(result: dict[str, Any] | None) -> bool:
    if not isinstance(result, dict):
        return False
    raw = (result.get("name") or "").strip()
    if not raw:
        return False
    lowered = raw.lower().strip()
    if lowered in _INVALID_NAMES:
        return False
    if lowered.startswith("unknown"):
        return False
    return True


async def lookup_go_upc(barcode: str) -> dict[str, Any] | None:
    settings = get_settings()
    api_key = (settings.go_upc_api_key or "").strip()
    if not api_key:
        return None

    b = (barcode or "").strip()
    if not b:
        return None

    url = f"https://go-upc.com/api/v1/code/{b}"
    timeout = httpx.Timeout(4.0, connect=2.0)
    async with httpx.AsyncClient(timeout=timeout, follow_redirects=True) as client:
        res = await client.get(
            url,
            params={"format": "true"},
            headers={"Authorization": f"Bearer {api_key}", "User-Agent": "main-inventory-app/1.0"},
        )

    if res.status_code != 200:
        return None

    data = res.json() if res.content else {}
    if not isinstance(data, dict):
        return None

    product = data.get("product")
    if not isinstance(product, dict):
        return None

    out: dict[str, Any] = {
        "barcode": b,
        "name": (product.get("name") or "").strip(),
        "brand": (product.get("brand") or "").strip() or None,
        "category": (product.get("category") or "").strip() or None,
        "image_url": (product.get("imageUrl") or "").strip() or None,
    }
    return out if is_valid_result(out) else None


async def lookup_upcitemdb(barcode: str) -> dict[str, Any] | None:
    settings = get_settings()
    b = (barcode or "").strip()
    if not b:
        return None

    user_key = (settings.upcitemdb_user_key or "").strip()
    key_type = (settings.upcitemdb_key_type or "3scale").strip() or "3scale"
    if user_key:
        url = "https://api.upcitemdb.com/prod/v1/lookup"
        headers = {"user_key": user_key, "key_type": key_type, "User-Agent": "main-inventory-app/1.0"}
    else:
        url = "https://api.upcitemdb.com/prod/trial/lookup"
        headers = {"User-Agent": "main-inventory-app/1.0"}

    timeout = httpx.Timeout(4.0, connect=2.0)
    async with httpx.AsyncClient(timeout=timeout, follow_redirects=True) as client:
        res = await client.get(url, params={"upc": b}, headers=headers)

    if res.status_code != 200:
        return None

    data = res.json() if res.content else {}
    if not isinstance(data, dict):
        return None

    items = data.get("items")
    if not isinstance(items, list) or not items:
        return None
    first = items[0]
    if not isinstance(first, dict):
        return None

    images = first.get("images")
    image_url = None
    if isinstance(images, list) and images:
        image_url = str(images[0]).strip() or None

    out: dict[str, Any] = {
        "barcode": b,
        "name": (first.get("title") or "").strip(),
        "brand": (first.get("brand") or "").strip() or None,
        "category": (first.get("category") or "").strip() or None,
        "image_url": image_url,
    }
    return out if is_valid_result(out) else None


async def lookup_openfoodfacts(barcode: str) -> dict[str, Any] | None:
    b = (barcode or "").strip()
    if not b:
        return None

    url = f"https://world.openfoodfacts.org/api/v0/product/{b}.json"
    timeout = httpx.Timeout(4.0, connect=2.0)
    async with httpx.AsyncClient(timeout=timeout, follow_redirects=True) as client:
        res = await client.get(url, headers={"User-Agent": "main-inventory-app/1.0"})

    if res.status_code != 200:
        return None

    data = res.json() if res.content else {}
    if not isinstance(data, dict) or data.get("status") != 1:
        return None

    product = data.get("product")
    if not isinstance(product, dict):
        return None

    name = (
        product.get("product_name")
        or product.get("product_name_en")
        or product.get("generic_name")
        or product.get("generic_name_en")
        or ""
    ).strip()
    brand = (product.get("brands") or "").strip() or None
    category = (product.get("categories") or "").strip() or None
    image_url = (product.get("image_url") or product.get("image_front_url") or "").strip() or None

    out: dict[str, Any] = {
        "barcode": b,
        "name": name,
        "brand": brand,
        "category": category,
        "image_url": image_url,
    }
    return out if is_valid_result(out) else None


class RenameDocumentRequest(BaseModel):
    storage_path: str
    display_name: str


class DocumentLinkRequest(BaseModel):
    storage_path: str
    item_id: str | None = None


@router.post("/add_item", response_model=AddItemResponse)
def add_item_route(payload: AddItemRequest, user: AuthenticatedUser = Depends(get_current_user)) -> AddItemResponse:
    limit_check = check_item_limit(user.user_id)
    if not limit_check["allowed"]:
        raise HTTPException(
            status_code=403,
            detail={
                "error": "item_limit_reached",
                "message": f"Free plan limit of {limit_check['limit']} items reached.",
                "current": limit_check["current"],
                "limit": limit_check["limit"],
                "upgrade_required": True,
            },
        )
    limits = check_free_tier_limits(user_id=user.user_id)
    if not limits["is_pro"]:
        if limits["at_item_limit"]:
            raise HTTPException(403, "FREE_TIER_ITEM_LIMIT")
        new_location = (payload.model_dump().get("location") or "Unsorted").strip()
        existing = search_items_basic(user_id=user.user_id, q="")
        existing_spaces = set(i.get("location", "Unsorted") for i in existing)
        if new_location not in existing_spaces and len(existing_spaces) >= 3:
            raise HTTPException(403, "FREE_TIER_SPACE_LIMIT")
    created = add_item(user_id=user.user_id, item=payload.model_dump())
    return AddItemResponse(item=created)


@router.post("/search_items", response_model=SearchItemsResponse)
def search_items_route(payload: SearchItemsRequest, user: AuthenticatedUser = Depends(get_current_user)) -> SearchItemsResponse:
    try:
        parsed = parse_search_query_to_keywords(query=payload.query)
        q = (parsed.get("text") or payload.query or "").strip()

        items = search_items_basic(user_id=user.user_id, q=q)

        category = parsed.get("category")
        location = parsed.get("location")
        if category:
            items = [i for i in items if (i.get("category") or "").lower() == str(category).lower()]
        if location:
            items = [i for i in items if (i.get("location") or "").lower() == str(location).lower()]

        try:
            create_activity(
                user_id=user.user_id,
                summary=f"Searched inventory: {payload.query}",
                metadata={"type": "search_items", "query": payload.query, "parsed": parsed, "results": len(items)},
                            )
        except Exception:
            logger.exception("Failed to write search activity")

        return SearchItemsResponse(items=items, parsed=parsed)
    except httpx.HTTPError:
        logger.exception("Upstream error during /search_items")
        raise service_unavailable("Search temporarily unavailable. Please try again.")
    except Exception:
        logger.exception("OpenAI error during /search_items")
        raise bad_gateway("Search temporarily unavailable. Please try again.")


@router.delete("/delete_item", response_model=DeleteItemResponse)
def delete_item_route(item_id: str, user: AuthenticatedUser = Depends(get_current_user)) -> DeleteItemResponse:
    ok = delete_item(user_id=user.user_id, item_id=item_id)
    return DeleteItemResponse(deleted=ok)


@router.patch("/update_item", response_model=UpdateItemResponse)
def update_item_route(payload: UpdateItemRequest, user: AuthenticatedUser = Depends(get_current_user)) -> UpdateItemResponse:
    try:
        updates = payload.model_dump(exclude_none=True)
        item_id = str(updates.pop("item_id"))
        updated = update_item(user_id=user.user_id, item_id=item_id, updates=updates)
        if not updated:
            raise bad_request("No updates applied")
        return UpdateItemResponse(item=updated)
    except Exception:
        logger.exception("Unhandled error during /update_item")
        raise service_unavailable("Update temporarily unavailable. Please try again.")


@router.post("/extract_from_image", response_model=ExtractFromImageResponse)
async def extract_from_image_route(
    file: UploadFile = File(...),
    user: AuthenticatedUser = Depends(get_current_user),
) -> ExtractFromImageResponse:
    raw = await file.read()
    if not raw:
        raise bad_request("Empty file")

    limit_check = await check_limit(user.user_id, "photo_scan")
    if not limit_check["allowed"]:
        raise HTTPException(
            status_code=429,
            detail={
                "error": "limit_exceeded",
                "feature": "photo_scan",
                "feature_label": limit_check["feature_label"],
                "current": limit_check["current"],
                "limit": limit_check["limit"],
                "message": f"You've used all {limit_check['limit']} free photo scans this month.",
            },
        )
    await increment_usage(user.user_id, "photo_scan")

    stored = upload_image(user_id=user.user_id, filename=file.filename or "upload.png", content=raw)
    try:
        extracted = extract_item_from_image(filename=file.filename or "upload.png", image_bytes=raw)
    except (openai.APITimeoutError, openai.APIConnectionError) as exc:
        logger.error("Vision extraction timed out (file=%s, size=%d): %s", file.filename, len(raw), exc)
        raise bad_gateway("Analysis timed out — please try a clearer photo.")
    except Exception:
        logger.exception("Vision extraction failed (file=%s, size=%d)", file.filename, len(raw))
        raise bad_gateway("AI extraction temporarily unavailable. Please try again.")

    return ExtractFromImageResponse(extracted=extracted, image_url=stored.url)


@router.post("/inventory/extract_from_image", response_model=MultiExtractFromImageResponse)
async def inventory_extract_from_image_route(
    file: UploadFile = File(...),
    user: AuthenticatedUser = Depends(get_current_user),
) -> MultiExtractFromImageResponse:
    raw = await file.read()
    if not raw:
        raise bad_request("Empty file")

    limit_check = await check_limit(user.user_id, "photo_scan")
    if not limit_check["allowed"]:
        raise HTTPException(
            status_code=429,
            detail={
                "error": "limit_exceeded",
                "feature": "photo_scan",
                "feature_label": limit_check["feature_label"],
                "current": limit_check["current"],
                "limit": limit_check["limit"],
                "message": f"You've used all {limit_check['limit']} free photo scans this month.",
            },
        )
    await increment_usage(user.user_id, "photo_scan")

    filename = file.filename or "upload.png"
    raw, filename = _convert_to_jpeg(raw, filename)

    try:
        data = extract_items_from_image_multi(filename=filename, image_bytes=raw)
    except openai.BadRequestError as exc:
        logger.error("Vision extraction bad request (file=%s): %s", file.filename, exc)
        raise HTTPException(status_code=422, detail=f"Could not analyze image: {str(exc)}")
    except Exception as exc:
        logger.exception("Vision extraction failed (file=%s, size=%d): %s", file.filename, len(raw), exc)
        raise HTTPException(status_code=500, detail="Image analysis failed. Please try a clearer photo.")

    items = data.get("items") or []
    summary = data.get("summary") or {"total_detected": len(items), "categories": {}}

    if not isinstance(summary, dict):
        summary = {"total_detected": len(items), "categories": {}}

    if "total_detected" not in summary:
        summary["total_detected"] = len(items)
    if "categories" not in summary:
        summary["categories"] = {}

    try:
        create_activity(
            user_id=user.user_id,
            summary=f"Scanned image for inventory items ({len(items)} detected)",
            metadata={"type": "scan_image", "filename": file.filename, "total_detected": len(items)},
                    )
    except Exception:
        logger.exception("Failed to write scan activity")

    return MultiExtractFromImageResponse(items=items, summary=summary)


@router.post("/inventory/bulk_create", response_model=BulkCreateResponse)
def inventory_bulk_create_route(
    payload: BulkCreateRequest,
    user: AuthenticatedUser = Depends(get_current_user),
) -> BulkCreateResponse:
    items_to_insert = [i.model_dump() for i in payload.items]
    if not is_pro_user(user.user_id):
        try:
            count_result = get_supabase_admin().table("items").select("item_id", count="exact").eq("user_id", user.user_id).execute()
            current = count_result.count or 0
        except Exception:
            current = 0
        slots_remaining = FREE_ITEM_LIMIT - current
        if slots_remaining <= 0:
            raise HTTPException(
                status_code=403,
                detail={
                    "error": "item_limit_reached",
                    "message": "Free plan limit of 30 items reached.",
                    "upgrade_required": True,
                },
            )
        items_to_insert = items_to_insert[:slots_remaining]
    try:
        inserted, failures = bulk_create_items(user_id=user.user_id, items=items_to_insert)

        try:
            create_activity(
                user_id=user.user_id,
                summary=f"Saved {len(inserted)} scanned items to inventory",
                metadata={"type": "bulk_create", "inserted": len(inserted), "failures": len(failures)},
                            )
        except Exception:
            logger.exception("Failed to write bulk create activity")

        return BulkCreateResponse(inserted=inserted, failures=failures)
    except httpx.HTTPError:
        logger.exception("Upstream error during bulk create")
        raise service_unavailable("Bulk insert temporarily unavailable. Please try again.")
    except Exception:
        logger.exception("Unhandled error during bulk create")
        raise service_unavailable("Bulk insert temporarily unavailable. Please try again.")


@router.get("/usage/status")
async def get_usage_status_endpoint(user: AuthenticatedUser = Depends(get_current_user)):
    """Return current free-tier usage counts for the authenticated user."""
    try:
        return await get_all_usage(user.user_id)
    except Exception as e:
        return {
            "is_pro": False,
            "error": str(e),
            "items": {"current": 0, "limit": FREE_ITEM_LIMIT, "remaining": FREE_ITEM_LIMIT},
            "photo_scans": {"current": 0, "limit": FREE_SCAN_LIMIT, "remaining": FREE_SCAN_LIMIT},
        }


@router.get("/usage")
async def get_usage(user: AuthenticatedUser = Depends(get_current_user)):
    """Get all usage limits for current user"""
    usage = await get_all_usage(user.user_id)
    return usage


@router.post("/usage/check")
async def check_usage_limit(request: Request, user: AuthenticatedUser = Depends(get_current_user)):
    """Check if a specific feature is allowed"""
    body = await request.json()
    feature = body.get("feature")
    if not feature:
        raise HTTPException(status_code=400, detail="feature required")
    result = await check_limit(user.user_id, feature)
    return result


@router.post("/usage/increment")
async def increment_usage_count(request: Request, user: AuthenticatedUser = Depends(get_current_user)):
    """Increment usage count for a feature after successful use"""
    body = await request.json()
    feature = body.get("feature")
    if not feature:
        raise HTTPException(status_code=400, detail="feature required")
    new_count = await increment_usage(user.user_id, feature)
    return {"count": new_count, "feature": feature}


@router.post("/process_barcode", response_model=ProcessBarcodeResponse)
async def process_barcode_route(
    payload: ProcessBarcodeRequest,
    user: AuthenticatedUser = Depends(get_current_user),
) -> ProcessBarcodeResponse:
    limit_check = await check_limit(user.user_id, "barcode_scan")
    if not limit_check["allowed"]:
        raise HTTPException(
            status_code=429,
            detail={
                "error": "limit_exceeded",
                "feature": "barcode_scan",
                "feature_label": limit_check["feature_label"],
                "current": limit_check["current"],
                "limit": limit_check["limit"],
                "message": f"You've used all {limit_check['limit']} free barcode scans this month.",
            },
        )
    guess = interpret_barcode(barcode=payload.barcode)
    await increment_usage(user.user_id, "barcode_scan")
    return ProcessBarcodeResponse(result=guess)


@router.post("/barcode_lookup", response_model=BarcodeLookupResponse)
async def barcode_lookup_route(
    payload: BarcodeLookupRequest,
    user: AuthenticatedUser = Depends(get_current_user),
) -> BarcodeLookupResponse:
    barcode = (payload.barcode or "").strip()
    if not barcode:
        raise bad_request("Missing barcode")
    limit_check = await check_limit(user.user_id, "barcode_scan")
    if not limit_check["allowed"]:
        raise HTTPException(
            status_code=429,
            detail={
                "error": "limit_exceeded",
                "feature": "barcode_scan",
                "feature_label": limit_check["feature_label"],
                "current": limit_check["current"],
                "limit": limit_check["limit"],
                "message": f"You've used all {limit_check['limit']} free barcode scans this month.",
            },
        )

    out: dict[str, Any] | None = None

    # STEP 0: Check internal parts_catalog first
    catalog_result = lookup_in_catalog(barcode)
    if catalog_result:
        return BarcodeLookupResponse(
            barcode=barcode,
            name=(catalog_result.get("name") or "Unknown item"),
            brand=catalog_result.get("brand") or None,
            model=catalog_result.get("part_number") or None,
            category=catalog_result.get("category") or None,
            image_url=None,
        )

    # STEP 1: Try UPCitemDB API
    try:
        out = await lookup_upcitemdb(barcode)
        if is_valid_result(out):
            save_to_catalog(barcode=barcode, data=out, source="upcitemdb")
    except Exception:
        logger.exception("UPCitemDB lookup failed")
        out = None

    # STEP 2: Try Go-UPC API
    if out is None:
        try:
            out = await lookup_go_upc(barcode)
            if is_valid_result(out):
                save_to_catalog(barcode=barcode, data=out, source="go_upc")
        except Exception:
            logger.exception("Go-UPC lookup failed")
            out = None

    # STEP 3: Try Open Food Facts API
    if out is None:
        try:
            out = await lookup_openfoodfacts(barcode)
            if is_valid_result(out):
                save_to_catalog(barcode=barcode, data=out, source="openfoodfacts")
        except Exception:
            logger.exception("OpenFoodFacts barcode lookup failed")
            out = None

    # STEP 4: Final fallback to AI logic
    if out is None:
        try:
            ai_out = await anyio.to_thread.run_sync(lambda: interpret_barcode(barcode=barcode))
            if isinstance(ai_out, dict):
                ai_norm: dict[str, Any] = {
                    "barcode": barcode,
                    "name": (ai_out.get("name") or "").strip(),
                    "brand": (ai_out.get("brand") or "").strip() or None,
                    "category": (ai_out.get("category") or "").strip() or None,
                    "image_url": (ai_out.get("image_url") or "").strip() or None,
                    "model": (ai_out.get("model") or ai_out.get("part_number") or None),
                }
                if is_valid_result(ai_norm):
                    out = ai_norm
                    save_to_catalog(barcode=barcode, data=out, source="ai")
        except Exception:
            logger.exception("AI barcode fallback failed")
            out = None

    if not is_valid_result(out):
        return BarcodeLookupResponse(
            barcode=barcode,
            name="Unknown item",
            brand=None,
            model=None,
            category=None,
            image_url=None,
        )

    await increment_usage(user.user_id, "barcode_scan")
    return BarcodeLookupResponse(
        barcode=barcode,
        name=(out.get("name") or "Unknown item"),
        brand=out.get("brand") or None,
        model=out.get("model") if isinstance(out, dict) else None,
        category=out.get("category") or None,
        image_url=out.get("image_url") or None,
    )


@router.post("/ai_command", response_model=AICommandResponse)
async def ai_command_route(
    payload: AICommandRequest,
    request: Request,
    user: AuthenticatedUser = Depends(get_current_user),
    stream: bool = False,
) -> AICommandResponse:
    limit_check = await check_limit(user.user_id, "ai_chat")
    if not limit_check["allowed"]:
        raise HTTPException(
            status_code=429,
            detail={
                "error": "limit_exceeded",
                "feature": "ai_chat",
                "feature_label": limit_check["feature_label"],
                "current": limit_check["current"],
                "limit": limit_check["limit"],
                "message": f"You've used all {limit_check['limit']} free AI chat messages this month.",
            },
        )

    accept = (request.headers.get("accept") or "").lower()
    wants_stream = bool(stream) or ("text/event-stream" in accept)

    if wants_stream:
        try:
            def _wrap_sse(gen):
                done_sent = False
                try:
                    for chunk in gen:
                        if not done_sent and isinstance(chunk, str) and '"type": "done"' in chunk:
                            done_sent = True
                        yield chunk
                except Exception:
                    logger.exception("AI command stream generator failed")
                finally:
                    if not done_sent:
                        yield 'event: end\n'
                        yield 'data: {"type":"done","tool":null,"result":null,"assistant_message":"Let me think about that..."}\n\n'

            gen = iter_ai_command_sse(user_id=user.user_id, message=payload.message, first_name=user.first_name, conversation_history=payload.conversation_history or None)
            wrapped = _wrap_sse(gen)
            await increment_usage(user.user_id, "ai_chat")
            return StreamingResponse(
                wrapped,
                media_type="text/event-stream",
                headers={
                    "Cache-Control": "no-cache",
                    "Connection": "keep-alive",
                    "X-Accel-Buffering": "no",
                },
            )
        except Exception:
            logger.exception("AI command stream failed")
            raise bad_gateway("AI temporarily unavailable. Please try again.")

    try:
        out = run_ai_command(user_id=user.user_id, message=payload.message, first_name=user.first_name, conversation_history=payload.conversation_history or None)
    except Exception:
        logger.exception("AI command failed")
        raise bad_gateway("AI temporarily unavailable. Please try again.")

    await increment_usage(user.user_id, "ai_chat")

    try:
        create_activity(
            user_id=user.user_id,
            summary="Used Assist",
            metadata={"type": "ai_chat", "tool": out.get("tool"), "message": payload.message},
                    )
    except Exception:
        logger.exception("Failed to write ai_chat activity")

    return AICommandResponse(
        tool=out.get("tool"),
        result=out.get("result"),
        assistant_message=out.get("assistant_message") or "Let me think about that...",
    )


@router.post("/ai_upload")
def ai_upload_route(
    request: Request,
    file: UploadFile = File(...),
    user: AuthenticatedUser = Depends(get_current_user),
) -> StreamingResponse:
    accept = (request.headers.get("accept") or "").lower()
    wants_stream = "text/event-stream" in accept
    if not wants_stream:
        raise bad_request("Streaming required")

    try:
        raw = file.file.read() if file.file is not None else b""
    except Exception:
        logger.exception("Failed to read uploaded file")
        raise bad_request("Invalid upload")

    def _wrap_sse(gen):
        done_sent = False
        try:
            for chunk in gen:
                if not done_sent and isinstance(chunk, str) and '"type": "done"' in chunk:
                    done_sent = True
                yield chunk
        except Exception:
            logger.exception("AI upload stream generator failed")
        finally:
            if not done_sent:
                yield 'event: end\n'
                yield 'data: {"type":"done","tool":null,"result":null,"assistant_message":""}\n\n'

    gen = iter_assist_file_analysis_sse(
        filename=file.filename or "upload",
        mime_type=file.content_type,
        content=raw,
    )
    wrapped = _wrap_sse(gen)
    return StreamingResponse(
        wrapped,
        media_type="text/event-stream",
        headers={
            "Cache-Control": "no-cache",
            "Connection": "keep-alive",
            "X-Accel-Buffering": "no",
        },
    )


@router.post("/documents/upload", response_model=UploadDocumentResponse)
async def upload_document_route(
    file: UploadFile = File(...),
    item_id: str | None = Form(None),
    user: AuthenticatedUser = Depends(get_current_user),
) -> UploadDocumentResponse:
    raw = await file.read()
    if not raw:
        raise bad_request("Empty file")

    filename = file.filename or "upload"
    content_type = (file.content_type or "").lower()
    allowed = {
        "application/pdf",
        "text/plain",
        "image/png",
        "image/jpg",
        "image/jpeg",
        "image/webp",
    }
    if content_type and content_type not in allowed:
        raise bad_request("Unsupported file type")

    try:
        stored = upload_document(user_id=user.user_id, filename=filename, content=raw)

        mime = (file.content_type or "").lower()
        file_type = "pdf" if (mime == "application/pdf" or filename.lower().endswith(".pdf")) else "image"

        doc = create_document(
            user_id=user.user_id,
            filename=filename,
            mime_type=file.content_type,
            storage_path=stored.path,
            file_type=file_type,
            size_bytes=len(raw),
            item_id=(item_id or "").strip() or None,
        )

        summary = summarize_activity(action="upload_document", details={"filename": filename, "mime_type": file.content_type})
        create_activity(user_id=user.user_id, summary=summary, metadata={"type": "upload_document", "storage_path": stored.path})

        return UploadDocumentResponse(document=doc, activity_summary=summary)
    except httpx.HTTPError:
        logger.exception("Upstream error during document upload")
        raise service_unavailable("Upload temporarily unavailable. Please try again.")
    except Exception:
        logger.exception("Unhandled error during document upload")
        raise service_unavailable("Upload temporarily unavailable. Please try again.")


@router.get("/documents", response_model=ListDocumentsResponse)
def list_documents_route(
    user: AuthenticatedUser = Depends(get_current_user),
    item_id: str | None = None,
    limit: int = 200,
) -> ListDocumentsResponse:
    docs = list_documents(user_id=user.user_id, limit=limit, item_id=item_id)
    return ListDocumentsResponse(documents=docs)


@router.patch("/documents/rename")
def rename_document_route(
    payload: RenameDocumentRequest,
    user: AuthenticatedUser = Depends(get_current_user),
) -> dict:
    storage_path = (payload.storage_path or "").strip()
    display_name = (payload.display_name or "").strip()
    if not storage_path:
        raise bad_request("Missing storage_path")
    if not display_name:
        raise bad_request("Missing display_name")

    try:
        doc = rename_document(user_id=user.user_id, storage_path=storage_path, display_name=display_name)
        if not doc:
            raise bad_request("Rename failed")
        return {"document": doc}
    except httpx.HTTPError:
        logger.exception("Upstream error during document rename")
        raise service_unavailable("Rename temporarily unavailable. Please try again.")
    except Exception:
        logger.exception("Unhandled error during document rename")
        raise service_unavailable("Rename temporarily unavailable. Please try again.")


@router.patch("/documents/link")
def link_document_route(
    payload: DocumentLinkRequest,
    user: AuthenticatedUser = Depends(get_current_user),
) -> dict:
    storage_path = (payload.storage_path or "").strip()
    if not storage_path:
        raise bad_request("Missing storage_path")

    try:
        doc = set_document_item_link(user_id=user.user_id, storage_path=storage_path, item_id=payload.item_id)
        if not doc:
            raise bad_request("Update failed")
        return {"document": doc}
    except httpx.HTTPError:
        logger.exception("Upstream error during document link update")
        raise service_unavailable("Update temporarily unavailable. Please try again.")
    except Exception:
        logger.exception("Unhandled error during document link update")
        raise service_unavailable("Update temporarily unavailable. Please try again.")


@router.delete("/documents", status_code=status.HTTP_204_NO_CONTENT)
def delete_document_route(
    storage_path: str,
    user: AuthenticatedUser = Depends(get_current_user),
) -> Response:
    if not storage_path or not storage_path.strip():
        raise bad_request("Missing storage_path")

    try:
        supabase = get_supabase_admin()
        supabase.storage.from_("documents").remove([storage_path])
        supabase.table("documents").delete().eq("user_id", user.user_id).eq("storage_path", storage_path).execute()
        return Response(status_code=status.HTTP_204_NO_CONTENT)
    except httpx.HTTPError:
        logger.exception("Upstream error during document deletion")
        raise service_unavailable("Delete temporarily unavailable. Please try again.")
    except Exception:
        logger.exception("Unhandled error during document deletion")
        raise service_unavailable("Delete temporarily unavailable. Please try again.")


@router.get("/activity/recent", response_model=RecentActivityResponse)
def recent_activity_route(
    user: AuthenticatedUser = Depends(get_current_user),
    limit: int = 10,
) -> RecentActivityResponse:
    try:
        activities = list_recent_activity(user_id=user.user_id, limit=limit)
        return RecentActivityResponse(activities=activities)
    except httpx.HTTPError:
        logger.exception("Upstream error during recent activity")
        raise service_unavailable("Activity temporarily unavailable. Please try again.")
    except Exception:
        logger.exception("Unhandled error during recent activity")
        raise service_unavailable("Activity temporarily unavailable. Please try again.")


@router.post("/sharing/create")
async def create_share_route(
    request: Request,
    user: AuthenticatedUser = Depends(get_current_user),
):
    body = await request.json()
    share_name = body.get("share_name", "My Inventory")
    permission = body.get("permission", "view")
    if permission not in ("view", "edit"):
        raise HTTPException(400, "Invalid permission")
    limit_check = await check_limit(user.user_id, "share_space")
    if not limit_check["allowed"]:
        raise HTTPException(
            status_code=429,
            detail={
                "error": "limit_exceeded",
                "feature": "share_space",
                "feature_label": limit_check["feature_label"],
                "current": limit_check["current"],
                "limit": limit_check["limit"],
                "message": f"Free plan allows {limit_check['limit']} active share. Upgrade for unlimited sharing.",
            },
        )
    result = sharing_service.create_share(
        user_id=user.user_id,
        share_name=share_name,
        permission=permission,
    )
    return result


@router.get("/sharing/my-shares")
def get_my_shares_route(
    user: AuthenticatedUser = Depends(get_current_user),
):
    return sharing_service.get_my_shares(user_id=user.user_id)


@router.post("/sharing/join")
async def join_share_route(
    request: Request,
    user: AuthenticatedUser = Depends(get_current_user),
):
    body = await request.json()
    share_code = (body.get("share_code") or "").strip().upper()
    if not share_code:
        raise HTTPException(400, "share_code is required")
    try:
        result = sharing_service.join_share(user_id=user.user_id, share_code=share_code)
        return result
    except ValueError as e:
        raise HTTPException(400, str(e))


@router.get("/sharing/joined")
def get_joined_shares_route(
    user: AuthenticatedUser = Depends(get_current_user),
):
    return sharing_service.get_joined_shares(user_id=user.user_id)


@router.delete("/sharing/{share_id}")
def revoke_share_route(
    share_id: str,
    user: AuthenticatedUser = Depends(get_current_user),
):
    sharing_service.revoke_share(user_id=user.user_id, share_id=share_id)
    return {"revoked": True}


@router.delete("/sharing/{share_id}/members/{member_id}")
def remove_member_route(
    share_id: str,
    member_id: str,
    user: AuthenticatedUser = Depends(get_current_user),
):
    try:
        sharing_service.remove_member(
            owner_user_id=user.user_id,
            share_id=share_id,
            member_user_id=member_id,
        )
        return {"removed": True}
    except ValueError as e:
        raise HTTPException(403, str(e))


@router.delete("/sharing/{share_id}/leave")
async def leave_share(
    share_id: str,
    user: AuthenticatedUser = Depends(get_current_user),
):
    client = get_supabase_admin()
    client.table("team_members").delete().eq(
        "share_id", share_id
    ).eq("member_user_id", user.user_id).execute()
    return {"left": True}


@router.get("/sharing/{share_id}/inventory")
def get_share_inventory_route(
    share_id: str,
    user: AuthenticatedUser = Depends(get_current_user),
):
    try:
        items = sharing_service.get_share_inventory(
            requesting_user_id=user.user_id,
            share_id=share_id,
        )
        return items
    except ValueError as e:
        raise HTTPException(403, str(e))


@router.post("/sharing/{share_id}/invite")
async def invite_member_by_email(
    share_id: str,
    body: dict,
    user: AuthenticatedUser = Depends(get_current_user),
):
    import resend
    import os
    resend.api_key = os.environ.get("RESEND_API_KEY", "")

    email = body.get("email", "").strip()
    if not email:
        raise HTTPException(400, "Email required")

    client = get_supabase_admin()
    share = client.table('team_shares').select('*').eq('share_id', share_id).eq('owner_user_id', user.user_id).execute()
    if not share.data:
        raise HTTPException(403, "Not your share")

    s = share.data[0]
    share_code = s['share_code']
    share_name = s.get('share_name') or 'a space'

    join_link = f"https://www.findez.ai/join?code={share_code}"

    try:
        resend.Emails.send({
            "from": "FindEZ <noreply@findez.ai>",
            "to": [email],
            "subject": f"You've been invited to view '{share_name}' on FindEZ",
            "html": f"""
            <div style="font-family: sans-serif; background: #000; color: #fff; padding: 40px; max-width: 500px; margin: auto; border-radius: 16px;">
              <h2 style="font-size: 22px; margin-bottom: 8px;">You're invited to FindEZ</h2>
              <p style="color: rgba(255,255,255,0.6); font-size: 15px;">Someone shared their inventory space <strong style="color:#fff">'{share_name}'</strong> with you.</p>
              <div style="margin: 32px 0; background: rgba(255,255,255,0.06); border-radius: 12px; padding: 20px; text-align: center;">
                <p style="color: rgba(255,255,255,0.5); font-size: 12px; letter-spacing: 2px; margin-bottom: 8px;">JOIN CODE</p>
                <p style="font-size: 32px; font-weight: 700; letter-spacing: 8px; color: #fff; margin: 0;">{share_code}</p>
              </div>
              <a href="{join_link}" style="display: block; background: #fff; color: #000; text-align: center; padding: 14px; border-radius: 99px; font-weight: 600; font-size: 15px; text-decoration: none;">Open FindEZ &amp; Join</a>
              <p style="color: rgba(255,255,255,0.3); font-size: 12px; margin-top: 24px; text-align: center;">Or enter code manually in the app: {share_code}</p>
            </div>
            """
        })
    except Exception as e:
        raise HTTPException(500, f"Email failed: {str(e)}")

    return {"sent": True, "email": email, "share_code": share_code}


@router.get("/sharing/{share_id}/members")
def get_share_members_route(
    share_id: str,
    user: AuthenticatedUser = Depends(get_current_user),
):
    try:
        return sharing_service.get_share_members(
            owner_user_id=user.user_id,
            share_id=share_id,
        )
    except ValueError as e:
        raise HTTPException(403, str(e))


def _is_date_like(s: str) -> bool:
    import re
    s = str(s).strip()
    # Matches patterns like 2023-03-04, 2023-03-04 00:00:00, etc.
    return bool(re.match(
        r'^\d{4}-\d{2}-\d{2}', s))


@router.post('/import/spreadsheet')
async def import_spreadsheet_route(
    file: UploadFile = File(...),
    location: str = Form('Unsorted'),
    user: AuthenticatedUser = Depends(get_current_user),
):
    import io as _io
    import csv as _csv
    import json as _json
    import re as _re
    import openpyxl

    raw = await file.read()
    if not raw:
        raise HTTPException(400, 'Empty file')

    filename = (file.filename or '').lower()
    all_sheets = []

    if filename.endswith('.xlsx') or filename.endswith('.xls'):
        try:
            wb = openpyxl.load_workbook(
                _io.BytesIO(raw), read_only=True, data_only=True)
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                rows = []
                for row in ws.iter_rows(values_only=True):
                    cleaned = [str(c).strip() if c is not None else '' for c in row]
                    if any(c for c in cleaned):
                        rows.append(cleaned)
                if len(rows) >= 2:
                    all_sheets.append({'name': sheet_name, 'headers': rows[0], 'rows': rows[1:]})
            wb.close()
        except Exception as e:
            raise HTTPException(422, f'Cannot read Excel: {e}')

    elif filename.endswith('.csv'):
        try:
            text = raw.decode('utf-8', errors='replace')
            reader = _csv.DictReader(_io.StringIO(text))
            headers = list(reader.fieldnames or [])
            rows = [[str(row.get(h, '')).strip() for h in headers] for row in reader]
            if rows:
                all_sheets.append({'name': 'Sheet1', 'headers': headers, 'rows': rows})
        except Exception as e:
            raise HTTPException(422, f'Cannot read CSV: {e}')
    else:
        raise HTTPException(400, 'Only .xlsx .xls .csv supported')

    if not all_sheets:
        raise HTTPException(422, 'No data found')

    # Build compact structure for AI — NOT all rows
    structure_lines = []
    for sheet in all_sheets:
        structure_lines.append(f"Sheet: {sheet['name']}")
        structure_lines.append(f"Columns: {', '.join(str(h) for h in sheet['headers'])}")
        structure_lines.append(f"Row count: {len(sheet['rows'])}")
        for i, row in enumerate(sheet['rows'][:3]):
            pairs = ', '.join(
                f'{sheet["headers"][j]}="{v}"'
                for j, v in enumerate(row)
                if v and j < len(sheet['headers'])
            )
            structure_lines.append(f"Sample {i+1}: {pairs}")
        structure_lines.append('')

    structure = '\n'.join(structure_lines)
    first_sheet = all_sheets[0]
    columns = ', '.join(str(h) for h in first_sheet['headers'])
    sample_rows_list = []
    for _sr_idx, _sr_row in enumerate(first_sheet['rows'][:3]):
        sample_rows_list.append({
            str(first_sheet['headers'][j]): _sr_row[j]
            for j in range(min(len(first_sheet['headers']), len(_sr_row)))
            if _sr_row[j]
        })
    sample_rows = _json.dumps(sample_rows_list, ensure_ascii=False)

    mapping_prompt = f"""You are mapping spreadsheet columns to inventory item fields.

Available inventory fields:
- name_columns: list of columns that form the item name/description (required)
- quantity_column: column with numeric quantity/count
- part_number_column: column with part numbers, SKUs, item codes, IDs (e.g. "PN-F177", "SKU", "Part #", "Item #", column named with a number like "8")
- subcategory_column: column with size, type, dimension, shaft size, screw length, thread size (e.g. "Shaft", "Screw Length", "Size", "M4", "Thread")
- brand_column: column with vendor name, supplier, manufacturer, brand (e.g. "Vendor Name", "Supplier", "Brand", "Manufacturer")
- purchase_source_column: column with vendor part numbers, supplier codes, order numbers
- notes_columns: list of columns with additional info, descriptions, specifications, comments

Spreadsheet structure:
Columns: {columns}
Sample rows: {sample_rows}

Rules:
- part_number_column: if a column name is just a number (like "8") or contains "part", "PN", "SKU", "ID", "code" -> map it here
- subcategory_column: shaft size, screw length, thread size, M2/M3/M4/M5/M6/M8 values -> map here
- brand_column: anything with "vendor", "supplier", "manufacturer", "brand" -> map here
- name_columns: use Description field if exists, otherwise combine meaningful text columns
- notes_columns: any remaining descriptive columns not mapped elsewhere
- Never use date columns as names
- quantity_column: only one column, must contain numbers

Return ONLY valid JSON, no markdown, no explanation:
{{
  "name_columns": ["col1"],
  "quantity_column": "col2",
  "part_number_column": "col3",
  "subcategory_column": "col4",
  "brand_column": "col5",
  "purchase_source_column": "col6",
  "notes_columns": ["col7"],
  "category": "Supplies",
  "display_columns": [
    {{"field": "name", "label": "Name"}},
    {{"field": "part_number", "label": "Part #"}},
    {{"field": "subcategory", "label": "Size/Type"}},
    {{"field": "brand", "label": "Vendor"}},
    {{"field": "purchase_source", "label": "Vendor Part #"}},
    {{"field": "quantity", "label": "Qty"}},
    {{"field": "notes", "label": "Notes"}}
  ]
}}
Only include fields in display_columns that actually have data in this spreadsheet.
Always include name and quantity."""

    settings = get_settings()
    ai_client = OpenAI(api_key=settings.openai_api_key)

    try:
        resp = ai_client.chat.completions.create(
            model='gpt-4o',
            max_completion_tokens=600,
            messages=[{'role': 'user', 'content': mapping_prompt}])
        mapping_raw = resp.choices[0].message.content.strip()
        if '```' in mapping_raw:
            for part in mapping_raw.split('```'):
                p = part.strip()
                if p.startswith('json'):
                    mapping_raw = p[4:].strip(); break
                elif p.startswith('{'):
                    mapping_raw = p; break
        mapping = _json.loads(mapping_raw)
    except Exception:
        mapping = {'name_columns': [], 'quantity_column': 'Quantity',
                   'part_number_column': None, 'subcategory_column': None,
                   'brand_column': None, 'purchase_source_column': None,
                   'notes_columns': [], 'category': 'Supplies'}

    def _get_val(row_dict: dict, col: str) -> str:
        if not col:
            return ''
        for k, v in row_dict.items():
            if str(k).strip().lower() == str(col).strip().lower():
                s = str(v).strip()
                return '' if s.lower() in ('none', 'nan', 'null', '') else s
        return ''

    items_to_insert = []
    for sheet in all_sheets:
        headers = sheet['headers']
        for row in sheet['rows']:
            row_dict = {
                headers[i]: row[i]
                for i in range(min(len(headers), len(row)))
            }

            name_columns = mapping.get('name_columns') or []
            if isinstance(name_columns, str):
                name_columns = [name_columns]

            notes_columns = mapping.get('notes_columns') or []
            if isinstance(notes_columns, str):
                notes_columns = [notes_columns]

            name = ' - '.join([
                _get_val(row_dict, col)
                for col in name_columns
                if _get_val(row_dict, col)
            ])
            if not name:
                continue

            qty_col = mapping.get('quantity_column') or ''
            qty_str = _get_val(row_dict, qty_col)
            try:
                qty = max(1, int(float(qty_str))) if qty_str else 1
            except Exception:
                qty = 1

            item = {
                'name': name[:500],
                'category': mapping.get('category') or 'Supplies',
                'subcategory': _get_val(row_dict, mapping.get('subcategory_column') or '') or None,
                'quantity': qty,
                'location': location,
                'part_number': _get_val(row_dict, mapping.get('part_number_column') or '') or None,
                'brand': _get_val(row_dict, mapping.get('brand_column') or '') or None,
                'purchase_source': _get_val(row_dict, mapping.get('purchase_source_column') or '') or None,
                'notes': ' | '.join([
                    _get_val(row_dict, col)
                    for col in notes_columns
                    if _get_val(row_dict, col)
                ]) or None,
            }
            items_to_insert.append(item)

    if not items_to_insert:
        raise HTTPException(422, 'No valid items found')

    limit_check = await check_limit(user.user_id, 'spreadsheet_import')
    if not limit_check['allowed']:
        raise HTTPException(
            status_code=429,
            detail={
                'error': 'limit_exceeded',
                'feature': 'spreadsheet_import',
                'feature_label': limit_check['feature_label'],
                'current': limit_check['current'],
                'limit': limit_check['limit'],
                'message': f"You've used all {limit_check['limit']} free spreadsheet imports this month.",
            },
        )
    await increment_usage(user.user_id, 'spreadsheet_import')

    inserted, failures = bulk_create_items(user_id=user.user_id, items=items_to_insert)

    try:
        create_activity(user_id=user.user_id,
            summary=f'Imported {len(inserted)} items from {file.filename}',
            metadata={'type': 'spreadsheet_import', 'filename': file.filename,
                      'inserted': len(inserted), 'failures': len(failures)})
    except Exception: pass

    return {
        'inserted': len(inserted),
        'failures': len(failures),
        'total_found': len(items_to_insert),
        'display_columns': mapping.get('display_columns', [
            {'field': 'name', 'label': 'Name'},
            {'field': 'quantity', 'label': 'Qty'},
            {'field': 'location', 'label': 'Location'},
        ]),
    }


@router.post("/checkouts/checkout")
async def checkout_item(
    body: dict,
    user: AuthenticatedUser = Depends(get_current_user),
):
    client = get_supabase_admin()
    item_id = body.get("item_id")
    checked_out_by = body.get("checked_out_by", "").strip()
    due_back_at = body.get("due_back_at")
    notes = body.get("notes")
    space_name = body.get("space_name", "").strip()

    if not item_id or not checked_out_by:
        raise HTTPException(400, "item_id and checked_out_by required")

    item = client.table("items").select("item_id, name, quantity").eq("item_id", item_id).eq("user_id", user.user_id).execute()
    if not item.data:
        raise HTTPException(404, "Item not found")

    result = client.table("checkouts").insert({
        "user_id": user.user_id,
        "item_id": item_id,
        "checked_out_by": checked_out_by,
        "due_back_at": due_back_at,
        "notes": notes,
        "space_name": space_name,
        "is_active": True,
    }).execute()

    return {"checkout": result.data[0] if result.data else {}}


@router.post("/checkouts/return")
async def return_item(
    body: dict,
    user: AuthenticatedUser = Depends(get_current_user),
):
    from datetime import datetime, timezone
    client = get_supabase_admin()
    checkout_id = body.get("checkout_id")
    if not checkout_id:
        raise HTTPException(400, "checkout_id required")

    # Fetch the checkout row first (no user_id filter — we check auth below)
    checkout_row = client.table("checkouts").select(
        "checkout_id, user_id, space_name"
    ).eq("checkout_id", checkout_id).execute()

    if not checkout_row.data:
        raise HTTPException(404, "Checkout not found")

    checkout = checkout_row.data[0]

    # Allow: the user who created the checkout, OR any owner/member of the space
    can_return = user.user_id == checkout.get("user_id")

    if not can_return:
        space_name = (checkout.get("space_name") or "").strip()
        if space_name:
            share_rows = client.table("team_shares").select(
                "share_id, owner_user_id"
            ).eq("share_name", space_name).execute()
            for s in (share_rows.data or []):
                if user.user_id == s["owner_user_id"]:
                    can_return = True
                    break
                member_check = client.table("team_members").select(
                    "member_id"
                ).eq("share_id", s["share_id"]).eq(
                    "member_user_id", user.user_id
                ).execute()
                if member_check.data:
                    can_return = True
                    break

    if not can_return:
        raise HTTPException(403, "Not authorized to return this checkout")

    client.table("checkouts").update({
        "returned_at": datetime.now(timezone.utc).isoformat(),
        "is_active": False,
    }).eq("checkout_id", checkout_id).execute()

    return {"returned": True}


@router.get("/checkouts/active")
async def get_active_checkouts(
    user: AuthenticatedUser = Depends(get_current_user),
):
    client = get_supabase_admin()

    # Get spaces the user owns
    owned_spaces = set()
    my_items = client.table("items").select("location").eq("user_id", user.user_id).execute()
    for item in (my_items.data or []):
        loc = (item.get("location") or "Unsorted").strip()
        if loc:
            owned_spaces.add(loc)

    # Get spaces the user has joined (via team_members)
    shared_spaces = set()
    joined = client.table("team_members").select(
        "team_shares(owner_user_id, share_name)"
    ).eq("member_user_id", user.user_id).execute()

    shared_owner_ids = set()
    for m in (joined.data or []):
        ts = m.get("team_shares") or {}
        owner_id = ts.get("owner_user_id")
        space_name = ts.get("share_name", "").strip()
        if owner_id:
            shared_owner_ids.add(owner_id)
        if space_name:
            shared_spaces.add(space_name.lower())

    # Also get spaces I share with others (I'm the owner)
    my_shares = client.table("team_shares").select(
        "share_name, share_id"
    ).eq("owner_user_id", user.user_id).eq("is_active", True).execute()

    shared_space_names_i_own = set()
    for s in (my_shares.data or []):
        sn = (s.get("share_name") or "").strip().lower()
        if sn:
            shared_space_names_i_own.add(sn)

    # Fetch my own checkouts
    my_checkouts = client.table("checkouts").select(
        "*, items(name, location, category)"
    ).eq("user_id", user.user_id).eq("is_active", True).order(
        "checked_out_at", desc=True
    ).execute()

    result = list(my_checkouts.data or [])

    # Fetch checkouts from teammates — only for shared spaces
    if shared_owner_ids:
        teammate_checkouts = client.table("checkouts").select(
            "*, items(name, location, category)"
        ).in_("user_id", list(shared_owner_ids)).eq("is_active", True).order(
            "checked_out_at", desc=True
        ).execute()

        for checkout in (teammate_checkouts.data or []):
            # Only include if the item is from a shared space
            item_data = checkout.get("items") or {}
            item_location = (item_data.get("location") or "").strip().lower()
            checkout_space = (checkout.get("space_name") or item_location).lower()

            if checkout_space in shared_spaces or item_location in shared_spaces:
                checkout["from_teammate"] = True
                result.append(checkout)

    # Also fetch checkouts from my team members for spaces I own and share
    if shared_space_names_i_own:
        my_share_ids = [s["share_id"] for s in (my_shares.data or [])]
        if my_share_ids:
            member_ids_result = client.table("team_members").select(
                "member_user_id"
            ).in_("share_id", my_share_ids).execute()

            member_ids = [m["member_user_id"] for m in (member_ids_result.data or [])]

            if member_ids:
                member_checkouts = client.table("checkouts").select(
                    "*, items(name, location, category)"
                ).in_("user_id", member_ids).eq("is_active", True).order(
                    "checked_out_at", desc=True
                ).execute()

                for checkout in (member_checkouts.data or []):
                    item_data = checkout.get("items") or {}
                    item_location = (item_data.get("location") or "").strip().lower()
                    checkout_space = (checkout.get("space_name") or item_location).lower()

                    if checkout_space in shared_space_names_i_own or item_location in shared_space_names_i_own:
                        checkout["from_teammate"] = True
                        if checkout not in result:
                            result.append(checkout)

    return {"checkouts": result}


@router.get("/checkouts/space")
async def get_space_checkouts(
    share_id: str,
    user: AuthenticatedUser = Depends(get_current_user),
):
    """Return active + recently returned checkouts strictly scoped to one shared space."""
    from datetime import datetime, timezone, timedelta

    client = get_supabase_admin()

    if not share_id:
        raise HTTPException(400, "share_id required")

    # Verify the share exists and the requesting user is owner or member
    share_rows = client.table("team_shares").select(
        "share_id, share_name, owner_user_id"
    ).eq("share_id", share_id).execute()

    if not share_rows.data:
        raise HTTPException(404, "Space not found")

    share = share_rows.data[0]
    share_name = share["share_name"]

    is_owner = user.user_id == share["owner_user_id"]
    is_member = False
    if not is_owner:
        m = client.table("team_members").select("member_id").eq(
            "share_id", share_id
        ).eq("member_user_id", user.user_id).execute()
        is_member = bool(m.data)

    if not is_owner and not is_member:
        raise HTTPException(403, "Not authorized")

    thirty_days_ago = (
        datetime.now(timezone.utc) - timedelta(days=30)
    ).isoformat()

    # Active checkouts for this space only
    active_result = client.table("checkouts").select(
        "*, items(name, location, category)"
    ).eq("space_name", share_name).eq("is_active", True).order(
        "checked_out_at", desc=True
    ).execute()

    # Checkouts returned within the last 30 days for this space
    returned_result = client.table("checkouts").select(
        "*, items(name, location, category)"
    ).eq("space_name", share_name).eq("is_active", False).gte(
        "returned_at", thirty_days_ago
    ).order("returned_at", desc=True).execute()

    return {
        "active": active_result.data or [],
        "returned": returned_result.data or [],
    }


@router.get("/checkouts/item/{item_id}")
async def get_item_checkouts(
    item_id: str,
    user: AuthenticatedUser = Depends(get_current_user),
):
    client = get_supabase_admin()

    visible_user_ids = {user.user_id}

    joined = client.table("team_members").select(
        "team_shares(owner_user_id)"
    ).eq("member_user_id", user.user_id).execute()

    for m in (joined.data or []):
        ts = m.get("team_shares") or {}
        owner_id = ts.get("owner_user_id")
        if owner_id:
            visible_user_ids.add(owner_id)

    owned = client.table("team_shares").select(
        "team_members(member_user_id)"
    ).eq("owner_user_id", user.user_id).eq("is_active", True).execute()

    for s in (owned.data or []):
        for m in (s.get("team_members") or []):
            member_id = m.get("member_user_id")
            if member_id:
                visible_user_ids.add(member_id)

    result = client.table("checkouts").select("*").eq(
        "item_id", item_id
    ).in_("user_id", list(visible_user_ids)).order(
        "checked_out_at", desc=True
    ).limit(10).execute()

    return {"checkouts": result.data or []}


@router.patch("/profile/update")
async def update_profile(
    body: dict,
    user: AuthenticatedUser = Depends(get_current_user),
):
    client = get_supabase_admin()
    updates = {}
    if "display_name" in body:
        updates["display_name"] = body["display_name"]
    if "contact_email" in body:
        updates["contact_email"] = body["contact_email"]
    if "avatar_color" in body:
        updates["avatar_color"] = body["avatar_color"]
    if not updates:
        raise HTTPException(400, "Nothing to update")
    client.table("profiles").upsert({"id": user.user_id, **updates}).execute()
    return {"updated": True}


@router.get("/profile/me")
async def get_my_profile(
    user: AuthenticatedUser = Depends(get_current_user),
):
    client = get_supabase_admin()
    profile = client.table("profiles").select("*").eq("id", user.user_id).execute()
    u = client.auth.admin.get_user_by_id(user.user_id)
    email = u.user.email if u and u.user else ""
    data = profile.data[0] if profile.data else {}
    return {
        "user_id": user.user_id,
        "email": email,
        "display_name": data.get("display_name") or email.split("@")[0],
        "contact_email": data.get("contact_email") or "",
        "avatar_color": data.get("avatar_color") or "#636366",
        "is_pro": data.get("is_pro", False),
    }
